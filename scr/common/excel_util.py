'''
-*- coding:utf-8 -*-
@Author:WillowZhang
@Email:116995292@qq.com
@Time:2021/10/20  17:58
'''

#xlrd(xls),openpyxl(xlsx)
import openpyxl


def read_excel(excel_url,sheet_name):
    '''读取excel'''
    # 加载工作簿
    wb = openpyxl.load_workbook(excel_url)
    # 获得sheet对象
    sheet = wb[sheet_name]
    # #打印最大行数和列数
    # print(sheet.max_row,sheet.max_column)
    # #取值（excel下标从1开始）
    # print(sheet.cell(1,1).value)

    all_list = []
    #循环行数
    for row in range(2,sheet.max_row+1):
        row_list = []
        # 循环列数
        for col in range(1,sheet.max_column+1):
            row_list.append(sheet.cell(row,col).value)
        all_list.append(row_list)
    # print(all_list[0][3])
    return all_list

# if __name__ == '__main__':
#     read_excel("../../data/data_login.xlsx","login")
